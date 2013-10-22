#!/usr/bin/env python
import os
import sys
import datetime
from twitter import Twitter, OAuth, oauth_dance, read_token_file
from models import TweetList
from openpyxl import  Workbook
from settings import *


def write_and_flush(s):
    sys.stdout.write(s); sys.stdout.flush()

print '\nTwitarian, version %s' % VERSION
print '<http://github.com/bcattle/twitarian/>'
print 'Bryan Cattle, (c) 2013\n'

# Ask the user for their screenname
while True:
    twitter_account = raw_input('Please enter your twitter username without the "@" sign: ')
    if twitter_account and twitter_account.strip()[0] != '@':
        break
twitter_account = twitter_account.strip()

# The output filename
OUTPUT_FILENAME     = '%s - %s' % (
    twitter_account,
    #datetime.datetime.now(tz=LOCAL_TIME).strftime('%b %d, %Y %I.%M.%S %p')
    datetime.datetime.now().strftime('%b %d, %Y %I.%M.%S %p')
)


# Try to guess what quarter the user is interested in
this_year = datetime.date.today().year
quarter_start_dates = [
    datetime.date(this_year, 1, 1),
    datetime.date(this_year, 4, 1),
    datetime.date(this_year, 7, 1),
    datetime.date(this_year, 10, 1),
]

# What was the start date of the most recently ended quarter?
last_qtr = filter(lambda x: x < datetime.date.today(), quarter_start_dates)[-1]

# Allow the user to enter a different date, if desired
print 'How far back do you want to go?'
print 'Press <enter> for the last quarter, which started on %s' % last_qtr.strftime('%B %d')
print 'Or enter a new date as [YYYY-MM-DD]'

while True:
    new_date = raw_input('>')

    if not new_date:
        # Accept the default
        start_date = datetime.datetime.combine(last_qtr, datetime.datetime.min.time())
        break
    else:
        # Try to parse the input
        try:
            start_date = datetime.datetime.strptime(new_date, '%Y-%m-%d')
            # It worked
            break
        except ValueError:
            print 'Sorry, couldn\'t understand the date "%s". Try again as [YYYY-MM-DD]' % new_date

print 'Okay, pulling your tweets back to %s' % start_date.strftime('%Y-%m-%d')

# Set the timezone so we can compare w/ what's returned by Twitter
start_date = UTC.localize(start_date)

# Auth
write_and_flush('Checking your password...')
if not os.path.exists(CREDENTIALS_FILE):
    oauth_dance('Twidarian', CONSUMER_KEY, CONSUMER_SECRET,
                CREDENTIALS_FILE)
oauth_token, oauth_secret = read_token_file(CREDENTIALS_FILE)
write_and_flush('password good.\n')


# Connect
write_and_flush('Connecting to twitter...')
t = Twitter(api_version=1.1,
            auth=OAuth(oauth_token, oauth_secret, CONSUMER_KEY, CONSUMER_SECRET))
write_and_flush('done\n')

# Pull tweets and mentions
write_and_flush('Pulling tweets...')
tweets = TweetList.pull_tweets_for_user(t, twitter_account, start_date)
write_and_flush('done\n')

write_and_flush('Pulling mentions...')
mentions = TweetList.pull_mentions(t, start_date)
write_and_flush('done\n')

# This isn't working for some reason. It's returning
# the same two results over and over again.
# However, the real search only returns 10 results, Oct-July
# so it's not that important.

#write_and_flush('Pulling retweets...')
#retweets = TweetList.pull_manual_retweets(t, TWITTER_ACCOUNT, START_DATE)
#write_and_flush('done\n')

# Save to a file
#filename = '%s.csv' % OUTPUT_FILENAME
#with open(filename, 'w') as csvfile:
#    tweets.save_output_file(csvfile)
#    mentions.save_output_file(csvfile)

# Save to an excel workbook
write_and_flush('Saving everything in an Excel file...')

wb = Workbook()
# First sheet, tweets
ws1 = wb.get_active_sheet()
ws1.title = 'Tweets'
tweets.save_into_worksheet(ws1)

# Second sheet, mentions
ws2 = wb.create_sheet(title='Mentions')
mentions.save_into_worksheet(ws2)

# Third sheet, retweets
ws3 = wb.create_sheet(title='Retweets')
#retweets.save_into_worksheet(ws3)

# Save the file
filename = '%s.xlsx' % OUTPUT_FILENAME
wb.save(filename)

write_and_flush('done!\n')


print '\nEverything ran successfully. The data was saved to the file'
print '\n\t%s\n' % filename

# If we're in Windows, offer to open the file in Excel
if os.name == 'nt':

    print '\nPress <enter> to open this file in Excel,'
    print 'otherwise press any other key to quit.'

    import msvcrt
    c = msvcrt.getch()
    if c == '\r':
        # Open in excel
        #os.system('start "%s\\%s"' % (sys.path[0], filename))
        from subprocess import Popen
        p = Popen(filename, shell=True)

else:
    raw_input('Press any key to continue')
    #os.system('read -s -n 1 -p "Press any key to continue..."')
    #print




# Calculate unique mentioners
#mention_usernames = []
#mention_user_by_username = {}
#for mention in mentions:
#    mention_usernames.append(mention.user_handle)
#    mention_user_by_username[mention.user_handle] = {
#        'followers':
#        'total_tweets':
#    }
#    self.user_followers     = raw_mention['user']['followers_count']
#    self.user_total_tweets  = raw_mention['user']['statuses_count']
#    self.location           = raw_mention['user']['location']
#    self.user_name          = raw_mention['user']['name']
#    self.user_handle
#
#mentioner_counter = collections.Counter()

