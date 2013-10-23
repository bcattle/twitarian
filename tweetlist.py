# -*- coding: utf-8 -*-
import collections
from models import Tweet, Mention, Mentioner
from twitter import TwitterHTTPError


class TweetList(list):
    """
    Results collection.
    Holds both tweets *and* mentions (which are also tweets)
    """

    def __init__(self, name, raw_tweets=None, tweet_klass=None,
                 processed_tweets=None):
        assert raw_tweets and tweet_klass or processed_tweets

        super(TweetList, self).__init__()
        self.name = name.capitalize()

        if raw_tweets:
            for raw_tweet in raw_tweets:
                self.append(tweet_klass(raw_tweet))
        else:
            self.extend(processed_tweets)
        self.tweets_by_day = None

    def get_tweets_by_day(self):
        if not self.tweets_by_day:
            # Bucket the tweets by day
            self.tweets_by_day = collections.defaultdict(list)
            for tweet in self:
                key = tweet.created_local.strftime('%m/%d/%y')
                self.tweets_by_day[key].append(tweet)
        return self.tweets_by_day

    @staticmethod
    def pull_tweets_for_user(twitter, screenname, start_date=None):
        """
        This can pull tweets for *any user*, assuming their tweets are public
        """
        last_id = None
        raw_tweets = []
        while True:
            params = {
                'count': 200,
                'screen_name': screenname,
            }
            if last_id:
                params['max_id'] = last_id

            new_raw_tweets = twitter.statuses.user_timeline(**params)           # <----
            if not new_raw_tweets:
                break
            raw_tweets.extend(new_raw_tweets)

            if start_date:
                # Is the date of our last tweet < start date?
                last_tweet = Tweet(raw_tweets[-1])
                if last_tweet.created_local > start_date:
                    last_id = raw_tweets[-1]['id_str']
                else:
                    break
            else:
                break

        tweet_list = TweetList(name='Tweets, chronological',
                               raw_tweets=raw_tweets, tweet_klass=Tweet)

        # If we only retrieved one set of results,
        # it may have overshot our `start_date`
        tweet_list = TweetList(
            name=tweet_list.name,
            processed_tweets=filter(
                lambda x: x.created_local > start_date, tweet_list
            )
        )
        # Sort by time created, descending
        tweet_list.sort(key=lambda x: x.created_local, reverse=True)
        return tweet_list


    @staticmethod
    def pull_mentions(twitter, start_date=None):
        """
        Only able to pull mentions for the **logged in user**
        """
        last_id = None
        raw_mentions = []
        while True:
            params = {
                'count': 200,
            }
            if last_id:
                params['max_id'] = last_id

            new_raw_mentions = twitter.statuses.mentions_timeline(**params)         # <----
            if not new_raw_mentions:
                break
            raw_mentions.extend(new_raw_mentions)

            if start_date:
                # Is the date of our last tweet < start date?
                last_tweet = Mention(raw_mentions[-1])
                if last_tweet.created_local > start_date:
                    last_id = raw_mentions[-1]['id_str']
                else:
                    break
            else:
                break

        mention_list = TweetList(name='Mentions',
                                 raw_tweets=raw_mentions, tweet_klass=Mention)

        # If we only retrieved one set of results,
        # it may have overshot our `start_date`
        mention_list = TweetList(
            name=mention_list.name,
            processed_tweets=filter(
                lambda x: x.created_local > start_date, mention_list
            )
        )
        # Sort by time created, descending
        mention_list.sort(key=lambda x: x.created_local, reverse=True)
        return mention_list


    @staticmethod
    def pull_manual_retweets(twitter, screenname, start_date=None):
        """
        Pulls "manual retweets" = those specified like "RT @comebody"
        The theory is that these aren't included in the official
        Twitter-returned "retweet count"

        --->> This function isn't that useful, because the search results are sparse
        Better data is obtained by filtering the "mentions" by "RT @username"
        """
        last_id = None
        raw_retweets = []
        while True:
            params = {
                'count': 200,
                'q': '"RT @%s"' % screenname
            }
            if last_id:
                params['max_id'] = last_id

            # This call returns the most recent tweets that were re-tweeted,
            # without specifying who actually did the retweeting
            #new_raw_retweets = twitter.statuses.retweets_of_me(**params)

            try:
                new_raw_retweets = twitter.search.tweets(**params)['statuses']         # <----

            except TwitterHTTPError, e:
                print 'Twitter returned an error, this probably means ' \
                      'we were rate-limited.'
                print '\tThe error was: %s' % e.response_data
                print '\tcontinuing with the data we have...'
                break

            if not new_raw_retweets:
                break
            raw_retweets.extend(new_raw_retweets)

            if start_date:
                # Is the date of our last tweet < start date?
                last_retweet = Mention(raw_retweets[-1])
                if last_retweet.created_local > start_date:
                    last_id = raw_retweets[-1]['id_str']
                else:
                    break
            else:
                break
        mention_list = TweetList(name='Re-tweets',
                                 raw_tweets=raw_retweets, tweet_klass=Mention)
        return mention_list


    def extract_retweets(self, username):
        """
        Returns a new TweetList consisting of mentions
        of the form "RT @username", pulled from the current TweetList
        """
        retweet_str = ('RT @%s' % username).lower()
        retweets = filter(lambda x: retweet_str in x.text.lower(), self)
        retweet_list = TweetList(name='Re-tweets',
                                 processed_tweets=retweets)
        return retweet_list


    def extract_most_retweeted(self):
        """
        Returns a new TweetList consisting of all tweets
        that were re-tweeted, in order of popularity
        """
        most_retweeted = filter(lambda x: x.retweets > 0, self)
        most_retweeted_list = TweetList(name='Tweets, most re-tweeted',
                                        processed_tweets=most_retweeted)
        most_retweeted_list.sort(key=lambda x: -x.retweets)
        return most_retweeted_list


    def extract_most_favorited(self):
        """
        Returns a new TweetList consisting of all tweets
        that were favorited, in order of popularity
        """
        most_favorited = filter(lambda x: x.favorites > 0, self)
        most_favorited_list = TweetList(name='Tweets, most favorited',
                                        processed_tweets=most_favorited)
        most_favorited_list.sort(key=lambda x: -x.favorites)
        return most_favorited_list


    def extract_unique_mentioners(self, username):
        """
        Returns a new TweetList with all the individuals
        who mentioned you, sorted by how many times they did
        and their number of followers
        """
        # Pull out the column of mentioning usernames and count it
        mention_counter = collections.Counter(
            [mention.user_handle for mention in self]
        )

        # Make a mapping from username to tweet, so we can get user's info
        mentioning_users_by_username = {
            mention.user_handle: mention for mention in self
        }

        # Make a list of /Mentioners/ from the above
        top_mentioners = [
            Mentioner(mention_count[1],
                      mentioning_users_by_username[mention_count[0]],
                      username)
            for mention_count in mention_counter.most_common()
        ]

        # Make the new TweetList
        mentioner_list = TweetList(name='Who mentioned you',
                                   processed_tweets=top_mentioners)

        # Already sorted by # mentiones, also want to sort by # followers
        mentioner_list.sort(key=lambda x: (-x.mention_count, -x.user_followers))

        return mentioner_list


    def save_output_file(self, username, file_object):
        """
        Saves to a flat CSV file, one table after the other
        """
        import csv
        writer = csv.writer(file_object)
        writer.writerow(['Tweets/mentions for @%s' % username])
        writer.writerow([])
        writer.writerow(self[0].to_dict().keys())
        for tweet in self:
            writer.writerow(tweet.to_dict().values())
        writer.writerow([])
        writer.writerow([])


    def save_into_worksheet(self, ws):
        """
        Saves output data to am Excel workbook using OpenPyXL
        http://pythonhosted.org/openpyxl/tutorial.html

        See also ---
           http://scienceoss.com/write-excel-files-with-python-using-xlwt/
        """
        from openpyxl.cell import get_column_letter
        row_index = 0
        if len(self):
            # First, write the column headers
            for index, key in enumerate(self[0].to_dict().keys()):
                ws.cell(row=0, column=index).value = key

            # Then, iterate through all the rows and write
            for row_index, row in enumerate(self):
                for col_index, cell_value in \
                        enumerate(row.to_dict().values()):
                    ws.cell(row=row_index + 1, column=col_index).value \
                        = cell_value

            # Set the column widths
            for col_index in range(len(self[0].to_dict())):
                # Get the max value in the column
                # Start with the name
                col_name = self[0].to_dict().keys()[col_index]
                col_max_chars = len(col_name)
                for row_index in range(len(self)):
                    cell_len = len(
                        unicode(
                            ws.cell(row=row_index + 1, column=col_index).value
                        ).encode('ascii', 'replace')
                    )
                    if cell_len > col_max_chars:
                        col_max_chars = cell_len

                # Set the width
                ws.column_dimensions[get_column_letter(col_index + 1)].width = col_max_chars * 1.1

        return row_index
