from datetime import datetime
from openpyxl.style import Color


def get_filename(username, extension=''):
    extension_str = '.%s' % extension if extension else ''
    return '%s - %s%s' % (
        username,
        datetime.now().strftime('%b %d, %Y %I.%M.%S %p'),
        extension_str
    )


def save_to_csv(username, tweet_lists):
    filename = get_filename(username, extension='csv')
    with open(filename, 'w') as csvfile:
        for tweet_list in tweet_lists:
            tweet_list.save_output_file(csvfile)
    return filename


def write_to_workbook(workboook, tweet_lists):
    for i, tweet_list in enumerate(tweet_lists):
        if i == 0:
            # First sheet, already there
            ws = workboook.get_active_sheet()
        else:
            # Second sheet, create it
            ws = workboook.create_sheet()

        ws.title = tweet_list.name
        tweet_list.save_into_worksheet(ws)


def write_engagement_report(workbook,
                            username, start_date, end_date, follower_count,
                            tweets_worksheet_name, mentions_worksheet_name):
    """
    This function generates the engagement report,
    summary statistics about all the retrieved data

    It includes the following data:
     - Start date and end date
     - Follower count
     - Total # tweets
     - Total mentions
     - Total favorites
     - Total re-tweets

     - Avg re-tweets per tweet
     - Engagement ratio

    The engagement ratio is defined as
        "Total # mentions, favorites, and re-tweets
         divided by total # tweets"
    """
    ws = workbook.create_sheet(title='Engagement report')

    show_year = start_date.year != end_date.year
    dates_str = '%s - %s' % (
        pretty_date(start_date, year=show_year),
        pretty_date(end_date, year=show_year)
    )

    # Write the title
    title_cell = ws.cell(row=0, column=0)
    title_cell.value = 'Engagement Report for @%s' % username
    title_cell.style.font.bold = True
    ws.cell(row=1, column=0).value = dates_str

    # Follower count
    ws.cell(row=3, column=0).value = 'Total followers on %s' % pretty_date(end_date, year=False)
    ws.cell(row=3, column=3).value = follower_count

    # Total # tweets
    ws.cell(row=4, column=0).value = 'Total tweets, %s' % dates_str
    total_tweets_cell = ws.cell(row=4, column=3)
    total_tweets_cell.value = "=COUNTA('Tweets, chronological'!A:A)-1"

    # Table header
    HEADER_ROW = 6
    ws.cell(row=HEADER_ROW, column=1).value = 'Total'
    ws.cell(row=HEADER_ROW, column=2).value = 'Avg. per tweet'

    # Mentions
    MENTIONS_ROW = 7
    ws.cell(row=MENTIONS_ROW, column=0).value = 'Mentions'
    # Total
    total_mentions_cell = ws.cell(row=MENTIONS_ROW, column=1)
    total_mentions_cell.value = "=COUNTA(Mentions!A:A)-1"
    # Average
    ws.cell(row=MENTIONS_ROW, column=2).value = "=%s/%s" % (
        total_mentions_cell.address, total_tweets_cell.address
    )

    # Favorites
    FAVORITES_ROW = 8
    ws.cell(row=FAVORITES_ROW, column=0).value = 'Favorites'
    # Total
    total_favorites_cell = ws.cell(row=FAVORITES_ROW, column=1)
    total_favorites_cell.value = "=SUM('Tweets, chronological'!E:E)"
    # Average
    ws.cell(row=FAVORITES_ROW, column=2).value = "=%s/%s" % (
        total_favorites_cell.address, total_tweets_cell.address
    )

    # total re-tweets
    RETWEETS_ROW = 9
    ws.cell(row=RETWEETS_ROW, column=0).value = 'Re-tweets'
    # Total
    total_retweets_cell = ws.cell(row=RETWEETS_ROW, column=1)
    total_retweets_cell.value = "=SUM('Tweets, chronological'!D:D)"
    # Average
    ws.cell(row=RETWEETS_ROW, column=2).value = "=%s/%s" % (
        total_retweets_cell.address, total_tweets_cell.address
    )

    # engagement ratio
    ENGAGEMENT_ROW = 11
    ws.cell(row=ENGAGEMENT_ROW, column=0).value = 'Engagement ratio'
    explanation_cell = ws.cell(row=ENGAGEMENT_ROW + 1, column=0)
    explanation_cell.value = \
        '"Total # mentions, favorites, and re-tweets / qtr ' \
        'divided by total # tweets / qtr"'
    explanation_cell.style.font.italic =True

    engagement_ratio_cell = ws.cell(row=ENGAGEMENT_ROW, column=3)
    engagement_ratio_cell.value = \
        "=(%s + %s + %s)/%s" \
        % (total_mentions_cell.address, total_favorites_cell.address,
           total_retweets_cell.address, total_tweets_cell.address)
    engagement_ratio_cell.style.font.color.index = Color.DARKBLUE


def pretty_date(d, year=True):
    format_str = '%b %d'
    if year:
        format_str += ' %Y'
    return d.strftime(format_str)