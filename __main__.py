#!/usr/bin/env python
import os
import sys
#import logging
from twitter import Twitter, OAuth, oauth_dance, read_token_file
from models import TweetList
from openpyxl import  Workbook
from settings import *

#logger = logging.getLogger(__name__)

def write_and_flush(s):
    sys.stdout.write(s); sys.stdout.flush()

print '\nTwitarian, version %s' % VERSION
print '<http://github.com/bcattle/twitarian/>'
print 'Bryan Cattle, (c) 2013\n'

# Auth
write_and_flush('Checking your password...')
if not os.path.exists(CREDENTIALS_FILE):
    oauth_dance('Twidarian', CONSUMER_KEY, CONSUMER_SECRET,
                CREDENTIALS_FILE)
oauth_token, oauth_secret = read_token_file(CREDENTIALS_FILE)
write_and_flush('password good.\n')


# Connect
write_and_flush('Connecting to twitter...')
t = Twitter(api_version=1.1,
            auth=OAuth(oauth_token, oauth_secret, CONSUMER_KEY, CONSUMER_SECRET))
write_and_flush('done\n')

# Pull tweets and mentions
write_and_flush('Pulling tweets...')
tweets = TweetList.pull_tweets_for_user(t, TWITTER_ACCOUNT, START_DATE)
write_and_flush('done\n')

write_and_flush('Pulling mentions...')
mentions = TweetList.pull_mentions(t, START_DATE)
write_and_flush('done\n')

# This isn't working for some reason. It's returning
# the same two results over and over again.
# However, the real search only returns 10 results, Oct-July
# so it's not that important. 

#write_and_flush('Pulling retweets...')
#retweets = TweetList.pull_manual_retweets(t, TWITTER_ACCOUNT, START_DATE)
#write_and_flush('done\n')

# Save to a file
#with open('%s.csv' % OUTPUT_FILENAME, 'w') as csvfile:
#    tweets.save_output_file(csvfile)
#    mentions.save_output_file(csvfile)

# Save to an excel workbook
write_and_flush('Saving everything in an Excel file...')

wb = Workbook()
# First sheet, tweets
ws1 = wb.get_active_sheet()
ws1.title = 'Tweets'
tweets.save_into_worksheet(ws1)

# Second sheet, mentions
ws2 = wb.create_sheet(title='Mentions')
mentions.save_into_worksheet(ws2)

# Third sheet, retweets
ws3 = wb.create_sheet(title='Retweets')
retweets.save_into_worksheet(ws3)

# Save the file
wb.save('%s.xlsx' % OUTPUT_FILENAME)

write_and_flush('done!\n')



# Calculate unique mentioners
#mention_usernames = []
#mention_user_by_username = {}
#for mention in mentions:
#    mention_usernames.append(mention.user_handle)
#    mention_user_by_username[mention.user_handle] = {
#        'followers':
#        'total_tweets':
#    }
#    self.user_followers     = raw_mention['user']['followers_count']
#    self.user_total_tweets  = raw_mention['user']['statuses_count']
#    self.location           = raw_mention['user']['location']
#    self.user_name          = raw_mention['user']['name']
#    self.user_handle
#
#mentioner_counter = collections.Counter()

